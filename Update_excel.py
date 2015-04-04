__author__ = 'shawn.wang'

import xlsxwriter
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.colors import BLUE
from openpyxl import load_workbook

def add_test_id2link_text(replace, begin_search = 'testfilter=', end_search= '&', link_text = "http://qc.carefusion.com/qcreporting/td/TheOne_TestCaseApprovalReport.asp?folder=&designer=&testfilter=6125&Submit=Submit"):
    begin_search = link_text.find(begin_search)+len(begin_search)
    end = link_text.find(end_search, begin_search)+len(end_search)-1
    return link_text[:begin_search]+replace+link_text[end:]



# print add_test_id2link_text("5072")
def add_hyperlink_2cell(wb, row_begin, row_end, col_begin, col_end):
    worksheet = wb.worksheets[0]
    for i in range(row_begin, row_end+1):
        for j in range(col_begin, col_end+1):
            cell = worksheet.cell(row=i, column=j)
            if str(cell.value) != "N":
                link = add_test_id2link_text(str(cell.value))
                cell.hyperlink = link
                cell.font = Font(BLUE)
                cell.font = Font(size=11)
                cell.font.color.rgb = '000000FF'
    return wb


# wb = load_workbook('C:\Users\shawn.wang\Desktop\pages\Testcases for 3rd WBV1.1.xlsx')
# ws = wb.worksheets[0]
#
# cell = ws.cell(row = 4, column = 3)
# cell.hyperlink = "https://bitbucket.org/openpyxl/openpyxl/issue/19/cell-hyperlink-request"
#
# print cell
# print cell.value
# print ws['A1'].value
file = 'C:\Users\shawn.wang\Desktop\pages\Testcases for 3rd WBV1.1.xlsx'
wb = load_workbook(file)
wb = add_hyperlink_2cell(wb, 2, 190, 3, 3)
wb = add_hyperlink_2cell(wb, 2, 161, 8, 8)
wb.save('C:\Users\shawn.wang\Desktop\pages\Testcases for 3rd WBV1.2.xlsx')